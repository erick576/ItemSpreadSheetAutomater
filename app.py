import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
keep_Adding = 1
row_counter = 2

while keep_Adding == 1:
    isAdd = int(input('Add Item?'))
    if isAdd != 1:
        break
    while isAdd == 1:
        transaction_id = int(input("What is the transaction id"))
        product_id = int(input('What is your product id?'))
        price = float(input('What is the price of the item?'))
        new_trans_id = sheet.cell(row_counter, 1)
        new_prod_id = sheet.cell(row_counter, 2)
        new_price = sheet.cell(row_counter, 3)
        new_trans_id.value = transaction_id
        new_prod_id.value = product_id
        new_price.value = price
        add_or_stop = int(input('Will you keep adding or will you stop?'))
        isAdd = add_or_stop
        keep_Adding = add_or_stop
        row_counter = row_counter + 1


values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4
                   )
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions.xlsx')


for row in range(2, sheet.max_row + 1):
    corrected_price = sheet.cell(row, 3).value * 1.13
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

wb.save('transactions.xlsx')