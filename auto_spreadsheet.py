import openpyxl as xl  #
from openpyxl.chart import BarChart, Reference  # Importing Methods to use for the excel sheet
from datetime import date  #

wb = xl.load_workbook('transactions.xlsx')  # Load the workbook that the data is going to be stored on
sheet = wb['Sheet1']  # Store the sheet in the excel file that will be used
keep_Adding = 'Yes'  # String variable to enter the loop
transaction_date = date.today()  # Will store the date that each item is added in
product_id = 0  # Default Id for each product will increase after every item

# Counts total amount of items already stored so it can continue at the right row
row_counter = int(input('How much items have you added so far?: ')) + 2

# This loop will allow you to keep adding items with their specific price for as long as you want
# Date and Product Id are automatically added into the row for each item added
while keep_Adding == "Yes" or keep_Adding == 'yes':
    isAdd = input('Add Item?')
    if not isAdd == 'Yes' and not isAdd == 'yes':
        break
    while isAdd == 'Yes' or isAdd == 'yes':
        product_id = product_id + 1
        price = float(input('What is the price of the item?'))
        new_trans_id = sheet.cell(row_counter, 1)
        new_prod_id = sheet.cell(row_counter, 2)
        new_price = sheet.cell(row_counter, 3)
        new_trans_id.value = transaction_date
        new_prod_id.value = product_id
        new_price.value = price
        add_or_stop = input('Will you keep adding or will you stop? (Yes or No)')
        isAdd = add_or_stop
        keep_Adding = add_or_stop
        row_counter = row_counter + 1

# Initialize the parameters for the graph
values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4
                   )
# Add the parameters and then create the bar graph representing the data
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

# Save the info onto the excel sheet
wb.save('transactions.xlsx')

# Loop around all the price values and then calculate and store the cost with tax
for row in range(2, sheet.max_row + 1):
    corrected_price = sheet.cell(row, 3).value * 1.13
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# Save the info on the excel sheet
wb.save('transactions.xlsx')
